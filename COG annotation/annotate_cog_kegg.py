"""
annotate_cog_kegg.py
====================
Annota proteïnes de Haloquadratum walsbyi amb:
  - Categoria COG i funció: primer des de cog_result_table.tsv (lookup local),
    amb fallback a la API de KEGG si no es troba.
  - KEGG pathway: via API de KEGG (organisme: hwa).

Input:
  - input_cog.xlsx       : Excel amb columnes 'UniProtKB code' i 'Locus Tag'
  - cog_result_table.tsv : Taula de referència COG (COG -> Cat + Annotation)

Output:
  - input_cog_annotated.xlsx : Excel original + columnes d'anotació

Ús:
  python annotate_cog_kegg.py

Dependències: pandas, requests, openpyxl
"""

import re
import time
import json
import os
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuració ────────────────────────────────────────────────────────────

INPUT_XLSX        = "input_cog.xlsx"
COG_TSV           = "cog_result_table.tsv"
OUTPUT_XLSX       = "input_cog_annotated.xlsx"
CACHE_FILE        = "annotation_cache.json"
KEGG_ORGANISM     = "hwa"          # Codi KEGG de Haloquadratum walsbyi DSM 16790
REQUEST_DELAY     = 0.4            # Segons entre crides a API (evita rate-limit)

# Mapa complet de les 25 categories COG (lletra -> descripció)
COG_CATEGORY_MAP = {
    "A": "RNA processing and modification",
    "B": "Chromatin structure and dynamics",
    "C": "Energy production and conversion",
    "D": "Cell cycle control, cell division, chromosome partitioning",
    "E": "Amino acid transport and metabolism",
    "F": "Nucleotide transport and metabolism",
    "G": "Carbohydrate transport and metabolism",
    "H": "Coenzyme transport and metabolism",
    "I": "Lipid transport and metabolism",
    "J": "Translation, ribosomal structure and biogenesis",
    "K": "Transcription",
    "L": "Replication, recombination and repair",
    "M": "Cell wall/membrane/envelope biogenesis",
    "N": "Cell motility",
    "O": "Post-translational modification, protein turnover, chaperones",
    "P": "Inorganic ion transport and metabolism",
    "Q": "Secondary metabolites biosynthesis, transport and catabolism",
    "R": "General function prediction only",
    "S": "Function unknown",
    "T": "Signal transduction mechanisms",
    "U": "Intracellular trafficking, secretion, and vesicular transport",
    "V": "Defense mechanisms",
    "W": "Extracellular structures",
    "X": "Mobilome: prophages, transposons",
    "Z": "Cytoskeleton",
}

# ─── Cache persistent ─────────────────────────────────────────────────────────

def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r") as f:
            return json.load(f)
    return {}

def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)

# ─── Càrrega de dades ─────────────────────────────────────────────────────────

def load_inputs():
    df = pd.read_excel(INPUT_XLSX, dtype=str)
    df.columns = df.columns.str.strip()
    # Neteja espais i caràcters no-breaking dels locus tags
    df["Locus Tag"] = df["Locus Tag"].str.strip().str.replace(r"\s+", "", regex=True)
    df["UniProtKB code"] = df["UniProtKB code"].str.strip()
    return df

def load_cog_tsv():
    tsv = pd.read_csv(COG_TSV, sep="\t", dtype=str)
    tsv.columns = tsv.columns.str.strip()
    # Retorna diccionari: COG_ID -> {'Cat': ..., 'Annotation': ...}
    return tsv.set_index("COG")[["Cat", "Annotation"]].to_dict(orient="index")

# ─── UniProt API: obté COG per UniProt ID ────────────────────────────────────

def get_cog_from_uniprot(uniprot_id, cache):
    cache_key = f"uniprot_{uniprot_id}"
    if cache_key in cache:
        return cache[cache_key]

    result = {"cog_id": None, "cog_source": None}
    try:
        url = f"https://rest.uniprot.org/uniprotkb/{uniprot_id}.json"
        resp = requests.get(url, timeout=15)
        time.sleep(REQUEST_DELAY)
        if resp.status_code != 200:
            cache[cache_key] = result
            return result

        data = resp.json()

        # Cerca referència a COG a dbReferences
        db_refs = data.get("uniProtKBCrossReferences", [])
        for ref in db_refs:
            if ref.get("database") == "COG":
                cog_id = ref.get("id", "")
                if re.match(r"COG\d+", cog_id):
                    result = {"cog_id": cog_id, "cog_source": "UniProt"}
                    cache[cache_key] = result
                    return result

        # Cerca COG en qualsevol camp de text lliure
        full_text = json.dumps(data)
        match = re.search(r"COG\d{4}", full_text)
        if match:
            result = {"cog_id": match.group(0), "cog_source": "UniProt-text"}
            cache[cache_key] = result
            return result

    except Exception as e:
        print(f"  [WARN] UniProt API error per {uniprot_id}: {e}")

    cache[cache_key] = result
    return result

# ─── KEGG API: info del gen (pathway + possible COG) ─────────────────────────

def get_kegg_gene_info(locus_tag, cache):
    cache_key = f"kegg_gene_{locus_tag}"
    if cache_key in cache:
        return cache[cache_key]

    result = {"pathways": [], "cog_id": None}
    kegg_id = f"{KEGG_ORGANISM}:{locus_tag}"

    try:
        url = f"https://rest.kegg.jp/get/{kegg_id}"
        resp = requests.get(url, timeout=15)
        time.sleep(REQUEST_DELAY)

        if resp.status_code != 200 or not resp.text.strip():
            cache[cache_key] = result
            return result

        text = resp.text

        # Extreu pathways
        in_pathway = False
        pathways = []
        for line in text.split("\n"):
            if line.startswith("PATHWAY"):
                in_pathway = True
                content = line[12:].strip()
                if content:
                    pathways.append(content)
            elif in_pathway and line.startswith(" "):
                content = line.strip()
                if content:
                    pathways.append(content)
            elif in_pathway:
                in_pathway = False

        # Cerca COG en DBLINKS, ORTHOLOGY o qualsevol camp
        cog_match = re.search(r"COG\d{4}", text)
        if cog_match:
            result["cog_id"] = cog_match.group(0)

        result["pathways"] = pathways

    except Exception as e:
        print(f"  [WARN] KEGG API error per {locus_tag}: {e}")

    cache[cache_key] = result
    return result

# ─── Anotació COG: TSV local + fallback KEGG ─────────────────────────────────

def annotate_cog(uniprot_id, locus_tag, cog_lookup, kegg_info, cache):
    """
    Estratègia de prioritat:
    1. UniProt API -> COG ID -> lookup al TSV local
    2. KEGG gene info -> COG ID -> lookup al TSV local
    3. No trobat
    """
    up_result = get_cog_from_uniprot(uniprot_id, cache)
    cog_id    = up_result.get("cog_id")
    source    = up_result.get("cog_source", "")

    # Fallback: COG des de KEGG
    if not cog_id and kegg_info.get("cog_id"):
        cog_id = kegg_info["cog_id"]
        source = "KEGG"

    # Lookup al TSV local
    if cog_id and cog_id in cog_lookup:
        entry = cog_lookup[cog_id]
        cat_letters = entry["Cat"].strip()
        annotation  = entry["Annotation"]
        # Expandeix la primera lletra de categoria
        first_letter = cat_letters[0]
        cat_desc = COG_CATEGORY_MAP.get(first_letter, cat_letters)
        return {
            "COG ID"          : cog_id,
            "COG Category"    : cat_letters,
            "COG Description" : cat_desc,
            "COG Function"    : annotation,
            "COG Source"      : f"TSV ({source})",
        }

    # No trobat
    return {
        "COG ID"          : cog_id if cog_id else "Not found",
        "COG Category"    : "",
        "COG Description" : "",
        "COG Function"    : "",
        "COG Source"      : "Not found",
    }

# ─── Anotació KEGG pathway ───────────────────────────────────────────────────

def annotate_kegg_pathway(kegg_info):
    pathways = kegg_info.get("pathways", [])
    if not pathways:
        return {"KEGG Pathway IDs": "", "KEGG Pathway Names": "Not found"}

    ids   = []
    names = []
    for entry in pathways:
        parts = entry.split(None, 1)
        ids.append(parts[0])
        if len(parts) == 2:
            names.append(parts[1])

    return {
        "KEGG Pathway IDs"  : "; ".join(ids),
        "KEGG Pathway Names": "; ".join(names) if names else "Not found",
    }

# ─── Exportació Excel formatat ────────────────────────────────────────────────

def export_excel(df_result):
    df_result.to_excel(OUTPUT_XLSX, index=False, engine="openpyxl")
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active

    # Estils
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E4057")   # blau nit
    cog_fill    = PatternFill("solid", fgColor="D6EAF8")   # blau clar
    kegg_fill   = PatternFill("solid", fgColor="D5F5E3")   # verd clar
    orig_fill   = PatternFill("solid", fgColor="FDFEFE")   # blanc

    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cog_cols  = {"COG ID", "COG Category", "COG Description",
                 "COG Function", "COG Source"}
    kegg_cols = {"KEGG Pathway IDs", "KEGG Pathway Names"}

    col_fills = {}
    for i, col in enumerate(df_result.columns, 1):
        if col in cog_cols:
            col_fills[i] = cog_fill
        elif col in kegg_cols:
            col_fills[i] = kegg_fill
        else:
            col_fills[i] = orig_fill

    # Capçalera
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 32

    # Files de dades
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill      = col_fills.get(cell.column, orig_fill)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = border

    # Amplades de columna
    col_widths = {
        "UniProtKB code"    : 16,
        "Locus Tag"         : 14,
        "Molecular Function": 36,
        "Biological Process": 36,
        "Cellular Component": 30,
        "COG ID"            : 12,
        "COG Category"      : 14,
        "COG Description"   : 42,
        "COG Function"      : 42,
        "COG Source"        : 16,
        "KEGG Pathway IDs"  : 26,
        "KEGG Pathway Names": 52,
    }
    for col_idx, col_name in enumerate(df_result.columns, 1):
        width = col_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_XLSX)
    print(f"\n✓ Excel guardat: {OUTPUT_XLSX}")

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Anotació COG + KEGG  |  Haloquadratum walsbyi")
    print("=" * 60)

    print("\n[1/4] Carregant arxius d'entrada...")
    df         = load_inputs()
    cog_lookup = load_cog_tsv()
    cache      = load_cache()
    print(f"  Proteïnes a anotar : {len(df)}")
    print(f"  Entrades COG al TSV: {len(cog_lookup)}")

    results = []

    print("\n[2/4] Consultant APIs (UniProt + KEGG)...")
    for idx, row in df.iterrows():
        uniprot_id = str(row.get("UniProtKB code", "")).strip()
        locus_tag  = str(row.get("Locus Tag", "")).strip()
        print(f"  [{idx+1:02d}/{len(df)}] {uniprot_id} | {locus_tag}")

        # Info KEGG del gen
        kegg_info = get_kegg_gene_info(locus_tag, cache)

        # Anotació COG
        cog_data  = annotate_cog(uniprot_id, locus_tag, cog_lookup, kegg_info, cache)

        # Anotació KEGG pathway
        path_data = annotate_kegg_pathway(kegg_info)

        record = {
            "UniProtKB code"    : uniprot_id,
            "Locus Tag"         : locus_tag,
            "Molecular Function": row.get("Molecular Function", ""),
            "Biological Process": row.get("Biological Process", ""),
            "Cellular Component": row.get("Cellular Component", ""),
            **cog_data,
            **path_data,
        }
        results.append(record)

        # Desa cache cada 5 proteïnes
        if (idx + 1) % 5 == 0:
            save_cache(cache)

    save_cache(cache)

    print("\n[3/4] Construint DataFrame de resultats...")
    df_result = pd.DataFrame(results)

    cog_found  = (df_result["COG ID"] != "Not found").sum()
    path_found = (df_result["KEGG Pathway Names"] != "Not found").sum()
    print(f"  COG assignat     : {cog_found}/{len(df_result)}")
    print(f"  Pathway assignat : {path_found}/{len(df_result)}")

    print("\n[4/4] Exportant Excel...")
    export_excel(df_result)

    # Resum de categories COG
    cats_found = df_result[df_result["COG Category"] != ""]["COG Category"]
    if not cats_found.empty:
        print("\n[Resum categories COG trobades]")
        for cat, count in cats_found.value_counts().items():
            desc = COG_CATEGORY_MAP.get(cat.strip()[0], cat)
            print(f"  {cat:8s} ({count}x)  {desc}")

    print("\nFet!")

if __name__ == "__main__":
    main()
